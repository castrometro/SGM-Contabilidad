# api/views.py
from rest_framework import viewsets, permissions
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.decorators import api_view, permission_classes, action
from rest_framework_simplejwt.views import TokenObtainPairView
from django.http import HttpResponseRedirect, HttpResponse
from django.conf import settings
import os
import redis
import json


from .permissions import IsGerenteOrSelfOrReadOnly

from .models import Cliente, Contrato, Usuario, Industria, Area, Servicio, ServicioCliente, AsignacionClienteUsuario

from .serializers import (
    ClienteSerializer, ContratoSerializer,
    UsuarioSerializer, IndustriaSerializer,
    AreaSerializer, ServicioSerializer, ServicioClienteSerializer,
    AsignacionClienteUsuarioSerializer, CustomTokenObtainPairSerializer,
    ServicioClienteMiniSerializer, AnalistaPerformanceSerializer,
    AnalistaDetalladoSerializer, DashboardDataSerializer, 
    EstadisticasAnalistaSerializer, ClienteSimpleSerializer,
    UsuarioSupervisorSerializer, UsuarioAnalistaSerializer
)
from .permissions import (
    IsAuthenticatedAndActive, IsGerente, IsSupervisor, IsAnalista,
    ClienteAccess, ContratoAccess
)
from django.db.models import Count, Q, Avg, Sum, F
from django.utils import timezone
from datetime import datetime, timedelta
from decimal import Decimal
from rest_framework.parsers import MultiPartParser, FormParser
import tempfile
import sys, os
# Asegurar que el repo raíz esté en sys.path para importar pythonlocales/parser.py
REPO_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..'))
if REPO_ROOT not in sys.path:
    sys.path.append(REPO_ROOT)
from api import parser as auxiliar_parser

class CustomTokenObtainPairView(TokenObtainPairView):
    serializer_class = CustomTokenObtainPairSerializer


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def ping(request):
    return Response({"pong": True, "usuario": str(request.user)})


class UsuarioViewSet(viewsets.ModelViewSet):
    queryset = Usuario.objects.all()
    serializer_class = UsuarioSerializer
    permission_classes = [IsAuthenticatedAndActive & (IsGerente | IsSupervisor)]

    def get_queryset(self):
        """Filtrar usuarios por tipo si se especifica en los parámetros."""
        queryset = super().get_queryset()
        tipo_usuario = self.request.query_params.get('tipo_usuario', None)
        if tipo_usuario:
            queryset = queryset.filter(tipo_usuario=tipo_usuario)
        return queryset

    @action(detail=False, methods=["get"], permission_classes=[IsAuthenticated], url_path="me")
    def me(self, request):
        """Devuelve los datos del usuario autenticado."""
        serializer = self.get_serializer(request.user)
        return Response(serializer.data)

    @action(detail=False, methods=["get"], permission_classes=[IsAuthenticated & IsSupervisor], url_path="mis-analistas")
    def mis_analistas(self, request):
        """Devuelve los analistas supervisados por el supervisor autenticado."""
        if request.user.tipo_usuario != 'supervisor':
            return Response({"error": "Solo los supervisores pueden acceder a esta información"}, status=403)
        
        # Usar el serializer específico para supervisores
        serializer = UsuarioSupervisorSerializer(request.user)
        return Response(serializer.data)

    @action(detail=False, methods=["get"], permission_classes=[IsAuthenticated & IsSupervisor], url_path="clientes-supervisados")
    def clientes_supervisados(self, request):
        """Devuelve todos los clientes asignados a los analistas supervisados."""
        if request.user.tipo_usuario != 'supervisor':
            return Response({"error": "Solo los supervisores pueden acceder a esta información"}, status=403)
        
        # Obtener asignaciones de clientes de analistas supervisados
        asignaciones = AsignacionClienteUsuario.get_clientes_por_supervisor(request.user)
        
        # Serializar los datos
        data = []
        for asignacion in asignaciones:
            data.append({
                'cliente': ClienteSimpleSerializer(asignacion.cliente).data,
                'analista': {
                    'id': asignacion.usuario.id,
                    'nombre': asignacion.usuario.nombre,
                    'apellido': asignacion.usuario.apellido,
                    'correo_bdo': asignacion.usuario.correo_bdo
                },
                'fecha_asignacion': asignacion.fecha_asignacion
            })
        
        return Response(data)

    @action(detail=True, methods=["post"], permission_classes=[IsAuthenticated & (IsGerente | IsSupervisor)], url_path="asignar-supervisor")
    def asignar_supervisor(self, request, pk=None):
        """Asigna un supervisor a un analista."""
        analista = self.get_object()
        supervisor_id = request.data.get('supervisor_id')
        
        if not supervisor_id:
            return Response({"error": "Se requiere supervisor_id"}, status=400)
        
        try:
            supervisor = Usuario.objects.get(id=supervisor_id, tipo_usuario='supervisor')
        except Usuario.DoesNotExist:
            return Response({"error": "Supervisor no encontrado"}, status=404)
        
        # Verificar que supervisor puede supervisar al analista (misma área)
        if not supervisor.puede_supervisar_a(analista):
            return Response({"error": "El supervisor y analista deben compartir al menos un área"}, status=400)
        
        analista.supervisor = supervisor
        analista.save()
        
        serializer = self.get_serializer(analista)
        return Response(serializer.data)
    

class ClienteViewSet(viewsets.ModelViewSet):
    queryset = Cliente.objects.all()
    serializer_class = ClienteSerializer
    permission_classes = [IsAuthenticatedAndActive, ClienteAccess]

    def get_queryset(self):
        qs = super().get_queryset()
        if self.request.user.tipo_usuario == 'analista':
            asignados = AsignacionClienteUsuario.objects.filter(
                usuario=self.request.user
            ).values_list('cliente_id', flat=True)
            return qs.filter(pk__in=asignados)
        return qs
    
    @action(detail=False, methods=["get"], url_path="asignados")
    def asignados(self, request):
        asignados = AsignacionClienteUsuario.objects.filter(usuario=request.user)
        ids = asignados.values_list("cliente_id", flat=True)
        clientes = Cliente.objects.filter(id__in=ids)
        serializer = self.get_serializer(clientes, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=["get"], url_path="servicios", permission_classes=[IsAuthenticated])
    def servicios(self, request, pk=None):
        cliente = self.get_object()
        contratos = Contrato.objects.filter(cliente=cliente)
        serializer = ContratoSerializer(contratos, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['get'], url_path='servicios-area', permission_classes=[IsAuthenticated])
    def servicios_area(self, request, pk=None):
        cliente = self.get_object()
        user_area_ids = request.user.areas.values_list("id", flat=True)

        # Obtener los IDs de servicios de esas áreas
        servicio_ids = Servicio.objects.filter(area_id__in=user_area_ids).values_list("id", flat=True)

        # Obtener los servicios contratados del cliente que pertenezcan a esas áreas
        servicios = ServicioCliente.objects.filter(
            cliente=cliente,
            servicio_id__in=servicio_ids
        ).select_related('servicio', 'servicio__area')

        serializer = ServicioClienteMiniSerializer(servicios, many=True)
        return Response(serializer.data)

        

class ContratoViewSet(viewsets.ModelViewSet):
    queryset = Contrato.objects.all()
    serializer_class = ContratoSerializer
    permission_classes = [IsAuthenticatedAndActive, ContratoAccess]

    def get_queryset(self):
        qs = super().get_queryset()
        if self.request.user.tipo_usuario == 'analista':
            asignados = AsignacionClienteUsuario.objects.filter(
                usuario=self.request.user
            ).values_list('cliente_id', flat=True)
            return qs.filter(cliente_id__in=asignados)
        return qs

class IndustriaViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = Industria.objects.all()
    serializer_class = IndustriaSerializer
    permission_classes = [IsAuthenticatedAndActive]

class AreaViewSet(viewsets.ModelViewSet):
    queryset = Area.objects.all()
    serializer_class = AreaSerializer
    permission_classes = [IsAuthenticatedAndActive & IsGerente]

class ServicioViewSet(viewsets.ModelViewSet):
    queryset = Servicio.objects.all()
    serializer_class = ServicioSerializer
    permission_classes = [IsAuthenticatedAndActive]


class ServicioClienteViewSet(viewsets.ModelViewSet):
    queryset = ServicioCliente.objects.all()
    serializer_class = ServicioClienteSerializer
    permission_classes = [IsAuthenticatedAndActive]

class AsignacionClienteUsuarioViewSet(viewsets.ModelViewSet):
    queryset = AsignacionClienteUsuario.objects.select_related('cliente','usuario')
    serializer_class = AsignacionClienteUsuarioSerializer
    permission_classes = [IsAuthenticated, IsGerenteOrSelfOrReadOnly]

    def get_queryset(self):
        user = self.request.user
        qs = super().get_queryset()
        # superusuario ve todo
        if user.is_superuser:
            return qs
        # gerente ve sólo las asignaciones de usuarios de sus mismas áreas
        if user.tipo_usuario == 'gerente':
            return qs.filter(usuario__areas__in=user.areas.all()).distinct()
        # supervisor ve sólo sus propias asignaciones
        if user.tipo_usuario == 'supervisor':
            return qs.filter(usuario=user)
        # analista ve sólo sus propias asignaciones, read-only
        return qs.filter(usuario=user)
        
    def perform_create(self, serializer):
        """Validar que el cliente no esté ya asignado a otro analista en las mismas áreas"""
        cliente = serializer.validated_data['cliente']
        usuario_asignado = serializer.validated_data['usuario']
        
        # Obtener las áreas del usuario que se está asignando
        areas_usuario = list(usuario_asignado.areas.all())
        
        # Verificar si ya existe algún analista asignado al cliente en las mismas áreas
        asignaciones_existentes = AsignacionClienteUsuario.objects.filter(
            cliente=cliente,
            usuario__areas__in=areas_usuario
        ).exclude(usuario=usuario_asignado).select_related('usuario').prefetch_related('usuario__areas')
        
        # Verificar conflictos por área
        for asignacion in asignaciones_existentes:
            areas_conflicto = set(areas_usuario) & set(asignacion.usuario.areas.all())
            if areas_conflicto:
                areas_nombres = ', '.join([area.nombre for area in areas_conflicto])
                from rest_framework.exceptions import ValidationError
                raise ValidationError({
                    'cliente': f'Este cliente ya está asignado a {asignacion.usuario.nombre} {asignacion.usuario.apellido} en el área: {areas_nombres}'
                })
        
        serializer.save()


# =============================================================================
# ❌ ENDPOINTS DESHABILITADOS - BI, Dashboard, Gerente
# =============================================================================
# Los siguientes endpoints han sido deshabilitados:
# - /api/bi-analistas/ - Performance de analistas
# - /api/dashboard/ - Dashboard ejecutivo
# - /api/analistas-detallado/ - Análisis detallado de analistas
# - /api/gerente/* - Todos los endpoints exclusivos de gerente
# - /api/cobranza/parse-auxiliar/ - Parser de auxiliar CxC
# 
# Código comentado para referencia futura si se necesita restaurar
# =============================================================================

# class AnalistaPerformanceViewSet(viewsets.ReadOnlyModelViewSet):
#     serializer_class = AnalistaPerformanceSerializer
#     permission_classes = [IsAuthenticatedAndActive & IsGerente]
#
#     def get_queryset(self):
#         gerente = self.request.user
#         areas = gerente.areas.all()
#         return (
#             Usuario.objects.filter(tipo_usuario='analista', areas__in=areas)
#             .distinct()
#             .annotate(
#                 clientes_asignados=Count('asignaciones', distinct=True),
#                 cierres_contabilidad=Count(
#                     'cierrecontabilidad',
#                     filter=Q(cierrecontabilidad__area__in=areas),
#                     distinct=True,
#                 ),
#             )
#         )
#
#
# class DashboardViewSet(viewsets.ReadOnlyModelViewSet):
#     # ... [código completo del dashboard comentado] ...
#     pass
#
# class AnalistasDetalladoViewSet(viewsets.ReadOnlyModelViewSet):
#     # ... [código completo de analistas detallado comentado] ...
#     pass
#
# @api_view(['POST'])
# @permission_classes([IsAuthenticated])
# def parse_auxiliar_cxc(request):
#     # ... [código completo del parser CxC comentado] ...
#     pass

# =============================================================================
# FIN ENDPOINTS DESHABILITADOS
# =============================================================================


@api_view(['GET'])
@permission_classes([IsAuthenticatedAndActive & IsGerente])
def clientes_disponibles(request, analista_id):
    """Obtener clientes que no tienen analista asignado en las áreas del gerente"""
    try:
        analista = Usuario.objects.get(id=analista_id, tipo_usuario='analista')
        gerente = request.user
        areas_gerente = gerente.areas.all()
        areas_analista = analista.areas.all()
        
        # Obtener IDs de áreas comunes entre gerente y analista
        areas_gerente_ids = list(areas_gerente.values_list('id', flat=True))
        areas_analista_ids = list(areas_analista.values_list('id', flat=True))
        areas_comunes_ids = list(set(areas_gerente_ids) & set(areas_analista_ids))
        
        if not areas_comunes_ids:
            # No hay áreas en común, no hay clientes disponibles
            return Response([])
        
        # Clientes que tienen servicios contratados en las áreas comunes
        from .models import ServicioCliente, Servicio
        clientes_con_servicios_area = Cliente.objects.filter(
            servicios_contratados__servicio__area_id__in=areas_comunes_ids
        ).distinct()
        
        # 1. Excluir clientes que YA ESTÁN asignados a este analista
        clientes_ya_asignados_a_analista = AsignacionClienteUsuario.objects.filter(
            usuario=analista
        ).values_list('cliente_id', flat=True)
        
        # 2. Excluir clientes que ya tienen analista asignado en alguna de las áreas comunes (pero de otros analistas)
        clientes_ocupados_en_areas = []
        for cliente in clientes_con_servicios_area.exclude(id__in=clientes_ya_asignados_a_analista):
            asignaciones_existentes = AsignacionClienteUsuario.objects.filter(
                cliente=cliente,
                usuario__areas__id__in=areas_comunes_ids
            ).exclude(usuario=analista).prefetch_related('usuario__areas')
            
            if asignaciones_existentes.exists():
                # Verificar si hay conflicto de áreas
                for asignacion in asignaciones_existentes:
                    areas_asignado_ids = list(asignacion.usuario.areas.values_list('id', flat=True))
                    if set(areas_comunes_ids) & set(areas_asignado_ids):
                        clientes_ocupados_en_areas.append(cliente.id)
                        break
        
        # Obtener lista final de clientes disponibles
        clientes_excluidos = list(clientes_ya_asignados_a_analista) + clientes_ocupados_en_areas
        disponibles = clientes_con_servicios_area.exclude(id__in=clientes_excluidos)
        
        serializer = ClienteSimpleSerializer(disponibles, many=True)
        return Response(serializer.data)
        
    except Usuario.DoesNotExist:
        return Response({'error': 'Analista no encontrado'}, status=404)


@api_view(['GET'])
@permission_classes([IsAuthenticatedAndActive & IsGerente])
def clientes_asignados(request, analista_id):
    """Obtener clientes asignados a un analista específico"""
    try:
        analista = Usuario.objects.get(id=analista_id, tipo_usuario='analista')
        
        asignaciones = AsignacionClienteUsuario.objects.filter(
            usuario=analista
        ).select_related('cliente')
        
        clientes = [asig.cliente for asig in asignaciones]
        
        serializer = ClienteSimpleSerializer(clientes, many=True)
        return Response(serializer.data)
        
    except Usuario.DoesNotExist:
        return Response({'error': 'Analista no encontrado'}, status=404)


@api_view(['DELETE'])
@permission_classes([IsAuthenticatedAndActive & IsGerente])
def remover_asignacion(request, analista_id, cliente_id):
    """Remover asignación cliente-analista"""
    try:
        asignacion = AsignacionClienteUsuario.objects.get(
            usuario_id=analista_id,
            cliente_id=cliente_id
        )
        asignacion.delete()
        return Response({'success': True})
        
    except AsignacionClienteUsuario.DoesNotExist:
        return Response({'error': 'Asignación no encontrada'}, status=404)

@api_view(['POST', 'GET'])
@permission_classes([IsAuthenticated])
def dashboard_streamlit_redirect(request, cliente_id=None):
    """
    Redirigir al dashboard de Streamlit con el cliente específico cargado
    
    Métodos soportados:
    - POST: {"cliente_id": 123}
    - GET: /api/dashboard-streamlit/123/
    """
    try:
        # Obtener cliente_id desde diferentes fuentes
        if request.method == 'POST':
            cliente_id = request.data.get('cliente_id', cliente_id)
        
        if not cliente_id:
            return Response({'error': 'cliente_id es requerido'}, status=400)
        
        # Verificar que el cliente existe y el usuario tiene acceso
        try:
            cliente = Cliente.objects.get(id=cliente_id)
        except Cliente.DoesNotExist:
            return Response({'error': 'Cliente no encontrado'}, status=404)
        
        # Verificar permisos según el tipo de usuario
        user = request.user
        if user.tipo_usuario == 'analista':
            # Los analistas solo pueden ver clientes asignados
            if not AsignacionClienteUsuario.objects.filter(
                usuario=user, cliente=cliente
            ).exists():
                return Response({
                    'error': 'No tiene permisos para ver este cliente'
                }, status=403)
        elif user.tipo_usuario not in ['gerente', 'supervisor']:
            return Response({
                'error': 'No tiene permisos para acceder al dashboard'
            }, status=403)
        
        # Construir URL de Streamlit
        streamlit_host = os.getenv('STREAMLIT_CONTA_HOST', 'localhost')
        streamlit_port = os.getenv('STREAMLIT_CONTA_PORT', '8502')
        
        # URL con parámetros para cargar el cliente automáticamente
        streamlit_url = f"http://{streamlit_host}:{streamlit_port}/?cliente_id={cliente_id}"
        
        if request.method == 'POST':
            # Para POST, devolver la URL en JSON
            return Response({
                'success': True,
                'streamlit_url': streamlit_url,
                'cliente': {
                    'id': cliente.id,
                    'nombre': cliente.nombre,
                    'razon_social': cliente.razon_social
                },
                'message': f'Dashboard de {cliente.nombre} listo para abrir'
            })
        else:
            # Para GET, redirigir directamente
            return HttpResponseRedirect(streamlit_url)
            
    except Exception as e:
        return Response({
            'error': f'Error al acceder al dashboard: {str(e)}'
        }, status=500)


## =============================================================================
# 🧾 CAPTURA MASIVA RINDE GASTOS
# =============================================================================

import redis
import json
import os
from django.http import JsonResponse

def get_redis_client_db1():
    """
    Obtiene cliente Redis para db1 usando la configuración de Django
    Con soporte UTF-8 completo
    """
    redis_password = os.environ.get('REDIS_PASSWORD', '')
    if redis_password:
        return redis.Redis(
            host='redis', 
            port=6379, 
            db=1, 
            password=redis_password, 
            decode_responses=True,
            encoding='utf-8',
            encoding_errors='strict'
        )
    else:
        return redis.Redis(
            host='redis', 
            port=6379, 
            db=1, 
            decode_responses=True,
            encoding='utf-8',
            encoding_errors='strict'
        )

def get_redis_client_db1_binary():
    """
    Obtiene cliente Redis para db1 para datos binarios (sin decode_responses)
    Con soporte UTF-8 para metadatos
    """
    redis_password = os.environ.get('REDIS_PASSWORD', '')
    if redis_password:
        return redis.Redis(
            host='redis', 
            port=6379, 
            db=1, 
            password=redis_password, 
            decode_responses=False,
            encoding='utf-8'
        )
    else:
        return redis.Redis(
            host='redis', 
            port=6379, 
            db=1, 
            decode_responses=False,
            encoding='utf-8'
        )

# Endpoints de captura masiva de gastos movidos a rindegastos app
# Los endpoints captura_masiva_gastos y estado_captura_gastos
# ahora están en rindegastos.views_procesamiento


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def descargar_resultado_gastos(request, task_id):
    """
    Descargar el archivo Excel procesado
    """
    try:
        redis_client = get_redis_client_db1()
        
        # Verificar que la tarea esté completada - incluir usuario_id en la clave
        metadata_raw = redis_client.get(f"captura_gastos_meta:{request.user.id}:{task_id}")
        if not metadata_raw:
            return Response({
                'error': 'No se encontró información de la tarea'
            }, status=404)
        
        metadata = json.loads(metadata_raw)
        
        if metadata.get('estado') != 'completado':
            return Response({
                'error': 'La tarea aún no ha sido completada'
            }, status=400)
        
        # Obtener el archivo Excel desde Redis (usar cliente binario) - incluir usuario_id en la clave
        redis_client_binary = get_redis_client_db1_binary()
        excel_content = redis_client_binary.get(f"captura_gastos_excel:{request.user.id}:{task_id}")
        if not excel_content:
            return Response({
                'error': 'El archivo procesado no está disponible'
            }, status=404)
        
        # Crear respuesta HTTP con el archivo
        from django.http import HttpResponse
        
        response = HttpResponse(
            excel_content,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="gastos_procesados_{task_id}.xlsx"'
        
        return response
        
    except Exception as e:
        return Response({
            'error': f'Error descargando archivo: {str(e)}'
        }, status=500)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def leer_headers_excel_gastos(request):
    """
    Endpoint para leer solo los headers de un archivo Excel
    También detecta automáticamente las posiciones de centros de costo
    """
    try:
        # Validar que se haya enviado un archivo
        if 'archivo' not in request.FILES:
            return Response({
                'error': 'No se encontró archivo en la petición'
            }, status=400)
        
        archivo = request.FILES['archivo']
        
        # Validar extensión del archivo
        if not archivo.name.lower().endswith(('.xlsx', '.xls')):
            return Response({
                'error': 'El archivo debe ser un Excel (.xlsx o .xls)'
            }, status=400)
        
        # Leer solo la primera fila para obtener headers
        from openpyxl import load_workbook
        from io import BytesIO
        
        archivo_content = archivo.read()
        workbook = load_workbook(BytesIO(archivo_content), read_only=True)
        sheet = workbook.active
        
        # Leer la primera fila (headers)
        headers = []
        primera_fila = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        
        for cell_value in primera_fila:
            if cell_value is not None:
                # Convertir a string con encoding UTF-8
                header_str = str(cell_value).encode('utf-8', errors='ignore').decode('utf-8')
                headers.append(header_str)
            else:
                headers.append('')
        
        # Detectar posiciones de centros de costo automáticamente
        centros_costo_detectados = {}
        for i, header in enumerate(headers):
            if header == 'PyC':
                centros_costo_detectados['PyC'] = {'posicion': i, 'nombre': header}
            elif header in ['PS', 'EB']:  # PS y EB son equivalentes
                centros_costo_detectados['PS'] = {'posicion': i, 'nombre': header}
            elif header == 'CO':
                centros_costo_detectados['CO'] = {'posicion': i, 'nombre': header}
            elif header == 'RE':
                centros_costo_detectados['RE'] = {'posicion': i, 'nombre': header}
            elif header == 'TR':
                centros_costo_detectados['TR'] = {'posicion': i, 'nombre': header}
            elif header == 'CF':
                centros_costo_detectados['CF'] = {'posicion': i, 'nombre': header}
            elif header == 'LRC':
                centros_costo_detectados['LRC'] = {'posicion': i, 'nombre': header}
        
        # Detectar posiciones de código y nombre de cuenta
        columnas_cuenta_detectadas = {}
        
        # Buscar primera columna que contenga "Codigo cuenta"
        for i, header in enumerate(headers):
            if header and 'Codigo cuenta' in str(header):
                columnas_cuenta_detectadas['codigo_cuenta'] = {'posicion': i, 'nombre': header}
                break
        
        # Buscar primera columna que contenga "Nombre cuenta"
        for i, header in enumerate(headers):
            if header and 'Nombre cuenta' in str(header):
                columnas_cuenta_detectadas['nombre_cuenta'] = {'posicion': i, 'nombre': header}
                break
        
        workbook.close()
        
        return Response({
            'headers': headers,
            'total_columnas': len(headers),
            'centros_costo': centros_costo_detectados,
            'columnas_cuenta': columnas_cuenta_detectadas,
            'mensaje': 'Headers leídos exitosamente'
        })
        
    except Exception as e:
        return Response({
            'error': f'Error leyendo headers del Excel: {str(e)}'
        }, status=500)