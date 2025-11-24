from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

import json
import unicodedata
from io import BytesIO

from django.http import HttpResponse
from openpyxl import load_workbook
from api.models import AsignacionClienteUsuario, ServicioCliente
from contabilidad.tasks import (
    get_headers_salida_contabilidad,
    get_redis_client_db1,
    get_redis_client_db1_binary,
)
## Endpoint sincrónico eliminado: se fuerza uso de Celery
from contabilidad.task_rindegastos import (
    reconstruir_excel_desde_json,
    rg_procesar_step1_task,
)
from rindegastos.models import Rendicion


def _extract_cliente_id(request):
    return (
        request.data.get("cliente_id")
        if hasattr(request, "data")
        else None
    ) or request.query_params.get("cliente_id")


def _user_has_rindegastos(user, cliente_id=None):
    """Valida que el usuario tenga algún cliente con el servicio RindeGastos."""
    base_qs = ServicioCliente.objects.filter(servicio__nombre__iexact="RindeGastos")

    if cliente_id:
        base_qs = base_qs.filter(cliente_id=cliente_id)

    assigned_clientes = AsignacionClienteUsuario.objects.filter(usuario=user).values_list("cliente_id", flat=True)

    if assigned_clientes.exists():
        base_qs = base_qs.filter(cliente_id__in=assigned_clientes)
    elif not user.is_staff and not user.is_superuser:
        return False

    return base_qs.exists()


def _normalize(text):
    if text is None:
        return ""
    if not isinstance(text, str):
        text = str(text)
    text = unicodedata.normalize('NFKD', text).encode('ascii', 'ignore').decode('ascii')
    return text.strip().lower()


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def leer_headers_excel_rindegastos(request):
    """
    Contabilidad: Endpoint exclusivo RindeGastos para leer headers y detectar CC
    como el rango entre la última columna 'Nombre cuenta' y la columna 'Fecha aprobacion'.
    """
    try:
        cliente_id = _extract_cliente_id(request)
        if not _user_has_rindegastos(request.user, cliente_id):
            return Response({'error': 'El servicio RindeGastos no está habilitado para este usuario/cliente'}, status=403)

        if 'archivo' not in request.FILES:
            return Response({'error': 'No se encontró archivo en la petición'}, status=400)

        archivo = request.FILES['archivo']

        if not archivo.name.lower().endswith(('.xlsx', '.xls')):
            return Response({'error': 'El archivo debe ser un Excel (.xlsx o .xls)'}, status=400)

        contenido = archivo.read()
        wb = load_workbook(BytesIO(contenido), read_only=True)
        ws = wb.active

        primera_fila = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = [(v if v is not None else '') for v in primera_fila]

        # Última columna que contiene "Nombre cuenta"
        last_nombre_idx = -1
        for i, h in enumerate(headers):
            if 'nombre cuenta' in _normalize(h):
                last_nombre_idx = i

        # Columna de "Fecha aprobacion"
        fecha_ap_idx = None
        for i, h in enumerate(headers):
            hn = _normalize(h)
            if 'fecha' in hn and 'aprobacion' in hn:
                fecha_ap_idx = i
                break

        centros_costo = {}
        if last_nombre_idx != -1 and fecha_ap_idx is not None and fecha_ap_idx - last_nombre_idx > 1:
            for pos in range(last_nombre_idx + 1, fecha_ap_idx):
                nombre = headers[pos]
                if nombre and str(nombre).strip() != '':
                    centros_costo[str(nombre)] = {"posicion": pos, "nombre": str(nombre)}

        # Fallback por nombres conocidos
        if not centros_costo:
            conocidos = ['PyC', 'PS', 'EB', 'CO', 'RE', 'TR', 'CF', 'LRC']
            for i, h in enumerate(headers):
                hs = str(h).strip() if h is not None else ''
                if hs in conocidos:
                    centros_costo[hs] = {"posicion": i, "nombre": hs}

        wb.close()

        return Response({
            'headers': [str(h) if h is not None else '' for h in headers],
            'total_columnas': len(headers),
            'centros_costo': centros_costo,
            'mensaje': 'Headers leídos exitosamente (RindeGastos/Contabilidad)'
        })

    except Exception as e:
        return Response({'error': f'Error leyendo headers del Excel: {str(e)}'}, status=500)


def _find_cc_range(headers):
    """Encuentra el rango (start_exclusive, end_inclusive_exclusive) para CC: entre última 'Nombre cuenta' y 'Fecha aprobacion'.
    Retorna (start_idx, end_idx). Si no hay coincidencia válida, retorna (None, None)."""
    last_nombre_idx = -1
    for i, h in enumerate(headers):
        if 'nombre cuenta' in _normalize(h):
            last_nombre_idx = i
    fecha_ap_idx = None
    for i, h in enumerate(headers):
        hn = _normalize(h)
        if 'fecha' in hn and 'aprobacion' in hn:
            fecha_ap_idx = i
            break
    if last_nombre_idx != -1 and fecha_ap_idx is not None and fecha_ap_idx - last_nombre_idx > 1:
        return last_nombre_idx + 1, fecha_ap_idx
    return None, None


def _parse_numeric(value):
    """Intenta parsear a float. Acepta strings con % o espacios. Retorna None si no parseable."""
    if value is None:
        return None
    try:
        if isinstance(value, str):
            s = value.replace('%', '').replace(',', '.').strip()
            if s == '':
                return None
            return float(s)
        if isinstance(value, (int, float)):
            return float(value)
    except Exception:
        return None
    return None


def _sanitize_sheet_name(name: str) -> str:
    s = str(name).replace(':', '-').replace('/', '-').replace('\\', '-')
    return s[:31] if len(s) > 31 else s


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def procesar_step1_rindegastos(request):
    """Inicia Step1 (asíncrono) exigiendo parametros_contables (JSON o campos sueltos)."""
    try:
        cliente_id = _extract_cliente_id(request)
        if not _user_has_rindegastos(request.user, cliente_id):
            return Response({'error': 'El servicio RindeGastos no está habilitado para este usuario/cliente'}, status=403)

        if 'archivo' not in request.FILES:
            return Response({'error': 'No se encontró archivo en la petición'}, status=400)
        archivo = request.FILES['archivo']
        if not archivo.name.lower().endswith(('.xlsx', '.xls')):
            return Response({'error': 'El archivo debe ser un Excel (.xlsx o .xls)'}, status=400)

        # Extraer cliente_servicio_id del request
        cliente_servicio_id = request.data.get('cliente_servicio_id')
        if not cliente_servicio_id:
            return Response({'error': 'Se requiere cliente_servicio_id'}, status=400)

        # Parse parametros contables obligatorios
        raw_param = request.data.get('parametros_contables')
        
        parametros_contables = None
        if raw_param:
            try:
                parametros_contables = json.loads(raw_param)
            except Exception as e:
                return Response({'error': 'parametros_contables no es JSON válido'}, status=400)
        else:
            # Fallback: construir desde campos sueltos
            cuenta_iva = request.data.get('cuentaIva') or request.data.get('cuenta_iva')
            cuenta_prov = request.data.get('cuentaProveedores') or request.data.get('cuenta_proveedores')
            cuenta_gasto = request.data.get('cuentaGasto') or request.data.get('cuenta_gasto')
            # CC: prefijo cc_*
            mapeo_cc = {}
            for k, v in request.data.items():
                if k.startswith('cc_') and v:
                    nombre = k[3:]
                    mapeo_cc[nombre] = v
            if any([cuenta_iva, cuenta_prov, cuenta_gasto]) or mapeo_cc:
                parametros_contables = {
                    'cuentasGlobales': {
                        'iva': cuenta_iva,
                        'proveedores': cuenta_prov,
                        'gasto_default': cuenta_gasto,
                    },
                    'mapeoCC': mapeo_cc
                }
        if not parametros_contables:
            return Response({'error': 'Se requieren parametros_contables (JSON) o campos individuales de cuentas y CC.'}, status=400)
        cg = parametros_contables.get('cuentasGlobales') or {}
        requeridas = ['iva', 'proveedores', 'gasto_default']
        faltantes = [r for r in requeridas if not cg.get(r)]
        if faltantes:
            return Response({'error': f'Faltan cuentasGlobales requeridas: {", ".join(faltantes)}'}, status=400)

        contenido = archivo.read()
        # Pasar cliente_servicio_id a la task
        task = rg_procesar_step1_task.delay(contenido, archivo.name, request.user.id, parametros_contables, int(cliente_servicio_id))
        return Response({
            'task_id': task.id,
            'estado': 'procesando',
            'archivo_nombre': archivo.name,
            'mensaje': 'Archivo enviado para Step1 (RG) con parametros contables'
        }, status=202)
    except Exception as e:
        return Response({'error': f'Error iniciando Step1: {str(e)}'}, status=500)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def estado_step1_rindegastos(request, task_id):
    try:
        cliente_id = _extract_cliente_id(request)
        if not _user_has_rindegastos(request.user, cliente_id):
            return Response({'error': 'El servicio RindeGastos no está habilitado para este usuario/cliente'}, status=403)

        r = get_redis_client_db1()
        meta_raw = r.get(f"rg_step1_meta:{request.user.id}:{task_id}")
        if not meta_raw:
            return Response({'error': 'No se encontró información de la tarea'}, status=404)
        return Response(json.loads(meta_raw))
    except Exception as e:
        return Response({'error': f'Error consultando estado: {str(e)}'}, status=500)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def descargar_step1_rindegastos(request, task_id):
    try:
        cliente_id = _extract_cliente_id(request)

        r = get_redis_client_db1()
        meta_raw = r.get(f"rg_step1_meta:{request.user.id}:{task_id}")
        meta = json.loads(meta_raw) if meta_raw else None

        # Validar permisos con el cliente solicitado o con la rendición asociada
        rendicion = Rendicion.objects.filter(datos_archivo__task_id=task_id).select_related('cliente_servicio__cliente').first()
        rendicion_cliente_id = getattr(getattr(rendicion, 'cliente_servicio', None), 'cliente_id', None)

        cliente_permiso = cliente_id or rendicion_cliente_id
        if cliente_permiso and not _user_has_rindegastos(request.user, cliente_permiso):
            return Response({'error': 'El servicio RindeGastos no está habilitado para este usuario/cliente'}, status=403)
        if not cliente_permiso and not _user_has_rindegastos(request.user, cliente_id):
            return Response({'error': 'El servicio RindeGastos no está habilitado para este usuario/cliente'}, status=403)

        if meta and meta.get('estado') != 'completado':
            return Response({'error': 'La tarea aún no ha sido completada'}, status=400)

        r_bin = get_redis_client_db1_binary()
        excel_content = r_bin.get(f"rg_step1_excel:{request.user.id}:{task_id}")

        # Si no está en Redis, intentar recrearlo desde la rendición guardada
        if not excel_content and rendicion and rendicion.datos_archivo:
            datos_archivo = rendicion.datos_archivo
            excel_json = datos_archivo.get('excel_json')
            if excel_json:
                excel_content = reconstruir_excel_desde_json(excel_json)
                if not meta:
                    meta = datos_archivo.get('metadata')
                # Rehidratar cache por 5 minutos para próximas descargas
                r_bin.setex(f"rg_step1_excel:{request.user.id}:{task_id}", 300, excel_content)
                if meta:
                    r.setex(f"rg_step1_meta:{request.user.id}:{task_id}", 300, json.dumps(meta, ensure_ascii=False))

        if not excel_content:
            return Response({'error': 'El archivo procesado no está disponible'}, status=404)

        nombre_archivo = 'rg_step1_{0}.xlsx'.format(task_id)
        if rendicion and rendicion.datos_archivo:
            nombre_archivo = rendicion.datos_archivo.get('archivo_nombre', nombre_archivo)

        resp = HttpResponse(
            excel_content,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        resp['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
        return resp
    except Exception as e:
        return Response({'error': f'Error descargando archivo: {str(e)}'}, status=500)
