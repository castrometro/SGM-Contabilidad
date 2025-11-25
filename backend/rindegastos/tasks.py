import json
from io import BytesIO

from celery import shared_task
from django.utils import timezone
from openpyxl import Workbook, load_workbook

from api.models import ServicioCliente, Usuario
from contabilidad.tasks import (
    get_headers_salida_contabilidad,
    get_redis_client_db1,
    get_redis_client_db1_binary,
)
from rindegastos.models import Rendicion, TipoDocumento


def _normalize(text):
    if text is None:
        return ""
    return str(text).strip()


def _parse_numeric(value):
    if value is None:
        return None
    try:
        if isinstance(value, str):
            s = value.replace('%', '').replace(',', '.').strip()
            if s == '':
                return None
            return float(s)
        if isinstance(value, (int, float)):
            return float(value)
    except Exception:
        return None
    return None


def _find_cc_range(headers):
    last_nombre_idx = -1
    for i, h in enumerate(headers):
        if 'nombre cuenta' in str(h).lower():
            last_nombre_idx = i
    fecha_ap_idx = None
    for i, h in enumerate(headers):
        hn = str(h).lower()
        if 'fecha' in hn and 'aprobacion' in hn:
            fecha_ap_idx = i
            break
    if last_nombre_idx != -1 and fecha_ap_idx is not None and fecha_ap_idx - last_nombre_idx > 1:
        return last_nombre_idx + 1, fecha_ap_idx
    return None, None


def reconstruir_excel_desde_json(excel_json):
    """Reconstruye un archivo Excel a partir de datos planos almacenados como JSON."""
    wb = Workbook()
    # Elimina la hoja por defecto para recrear las originales
    wb.remove(wb.active)

    sheets = excel_json.get('sheets', [])
    if not sheets:
        wb.create_sheet(title='Sheet')
        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    for sheet in sheets:
        ws = wb.create_sheet(title=sheet.get('nombre') or 'Sheet')
        headers = sheet.get('headers') or []
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)

        for row_idx, row_values in enumerate(sheet.get('rows', []), start=2):
            for col_idx, value in enumerate(row_values, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


@shared_task(bind=True)
def rg_procesar_step1_task(self, archivo_content, archivo_nombre, usuario_id, parametros_contables=None, cliente_servicio_id=None):
    """
    Genera Excel con hojas por grupo (Tipo Doc + cantidad de CC > 0) y guarda en Redis.
    Guarda metadatos en rg_step1_meta:{usuario_id}:{task_id} y el archivo en rg_step1_excel:{usuario_id}:{task_id}
    Crea un registro de Rendicion en la base de datos.
    """
    task_id = self.request.id
    redis_client = get_redis_client_db1()

    # Validar parametros contables obligatorios
    if not parametros_contables:
        raise ValueError("parametros_contables es obligatorio (cuentasGlobales y mapeoCC)")
    cuentas_globales = parametros_contables.get('cuentasGlobales') or {}
    
    mapeo_cc_param = parametros_contables.get('mapeoCC') or {}
    requeridas = ['iva', 'proveedores', 'gasto_default']
    faltantes = [r for r in requeridas if not cuentas_globales.get(r)]
    if faltantes:
        raise ValueError(f"Faltan cuentasGlobales requeridas: {', '.join(faltantes)}")

    # Crear registro de Rendicion si se proporcionó cliente_servicio_id
    rendicion = None
    if cliente_servicio_id:
        try:
            usuario = Usuario.objects.get(id=usuario_id)
            cliente_servicio = ServicioCliente.objects.get(id=cliente_servicio_id)

            rendicion = Rendicion.objects.create(
                cliente_servicio=cliente_servicio,
                usuario=usuario,
                fecha_ejecucion=timezone.now(),
                datos_archivo={
                    'archivo_nombre': archivo_nombre,
                    'task_id': task_id,
                    'parametros_contables': parametros_contables,
                }
            )
        except Exception as e:
            # Log error pero continuar con el procesamiento
            print(f"Error creando Rendicion: {str(e)}")

    # Mapeo código -> nombre de tipos de documento desde BD
    tipos_documento_map = {}
    if cliente_servicio_id:
        try:
            tipos_documento_map = {
                str(codigo).strip(): nombre
                for codigo, nombre in TipoDocumento.objects.filter(cliente_servicio_id=cliente_servicio_id).values_list('codigo', 'nombre')
            }
        except Exception as e:
            print(f"Error cargando tipos de documento para cliente_servicio {cliente_servicio_id}: {str(e)}")

    def tipo_doc_display(codigo):
        if codigo is None:
            return ""
        return tipos_documento_map.get(str(codigo).strip(), str(codigo))

    # Meta inicial
    metadata = {
        'task_id': task_id,
        'usuario_id': usuario_id,
        'archivo_nombre': archivo_nombre,
        'inicio': timezone.now().isoformat(),
        'estado': 'procesando',
        'grupos': [],
        'archivo_excel_disponible': False,
        'cuentas_globales_usadas': list(cuentas_globales.keys()),
        'rendicion_id': rendicion.id if rendicion else None,
    }
    redis_client.setex(
        f"rg_step1_meta:{usuario_id}:{task_id}", 300, json.dumps(metadata, ensure_ascii=False)
    )

    # Leer archivo y agrupar
    wb_in = load_workbook(BytesIO(archivo_content), read_only=True)
    ws_in = wb_in.active
    headers = [(v if v is not None else '') for v in next(ws_in.iter_rows(min_row=1, max_row=1, values_only=True))]

    # Índice Tipo Doc (robusto: sinónimos y case-insensitive)
    posibles_nombres = {'tipo doc', 'tipodoc', 'tipo_documento', 'tipo documento', 'tipo_doc'}
    idx_tipo_doc = None
    for i, h in enumerate(headers):
        nombre_norm = str(h).strip().lower()
        if nombre_norm in posibles_nombres:
            idx_tipo_doc = i
            break
    if idx_tipo_doc is None:
        raise ValueError("No se encontró la columna de Tipo de Documento (buscó: Tipo Doc / tipodoc / tipo_documento)")

    # Índice Monto Exento (solo para tipo doc 33) - Issue #174
    posibles_monto_exento = {'monto exento', 'montoexento', 'monto_exento', 'exento'}
    idx_monto_exento = None
    for i, h in enumerate(headers):
        nombre_norm = str(h).strip().lower()
        if nombre_norm in posibles_monto_exento:
            idx_monto_exento = i
            break

    # Índice Folio - Issues #3 y #4
    posibles_folio = {'folio', 'nro folio', 'numero folio', 'nrofolio'}
    idx_folio = None
    for i, h in enumerate(headers):
        nombre_norm = str(h).strip().lower()
        if nombre_norm in posibles_folio:
            idx_folio = i
            break

    # Índice RUT Proveedor - Para transferir a "Codigo Auxiliar"
    posibles_rut_proveedor = {'rut proveedor', 'rutproveedor', 'rut_proveedor', 'rut', 'codigo proveedor', 'codigo_proveedor'}
    idx_rut_proveedor = None
    for i, h in enumerate(headers):
        nombre_norm = str(h).strip().lower()
        if nombre_norm in posibles_rut_proveedor:
            idx_rut_proveedor = i
            break

    # Índice Fecha Docto - Para transferir a "Fecha Emisión" y "Fecha Vencimiento"
    posibles_fecha_docto = {'fecha docto', 'fechadocto', 'fecha_docto', 'fecha documento', 'fecha_documento', 'fecha doc', 'fechadoc'}
    idx_fecha_docto = None
    for i, h in enumerate(headers):
        nombre_norm = str(h).strip().lower()
        if nombre_norm in posibles_fecha_docto:
            idx_fecha_docto = i
            break

    cc_start, cc_end = _find_cc_range(headers)
    conocidos = ['PyC', 'PS', 'EB', 'CO', 'RE', 'TR', 'CF', 'LRC']
    cc_indices_conocidos = {str(h).strip(): i for i, h in enumerate(headers) if str(h).strip() in conocidos}

    grupos = {}
    grupos_filas = {}  # clave -> lista de (tipo_doc, cc_count, fila_index)
    total_filas = 0
    debug_filas = []
    # Guardar filas originales para cálculos posteriores
    filas_originales = {}
    for row_idx, row in enumerate(ws_in.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not any(row):
            continue
        total_filas += 1
        tipo_doc = row[idx_tipo_doc] if idx_tipo_doc < len(row) else None
        tipo_doc = str(tipo_doc) if tipo_doc is not None else 'Sin Tipo'
        tipo_doc_nombre = tipo_doc_display(tipo_doc) or 'Sin Tipo'

        # Contar CC válidos
        cc_count = 0
        if cc_start is not None and cc_end is not None:
            for col in range(cc_start, cc_end):
                val = row[col] if col < len(row) else None
                num = _parse_numeric(val)
                # Contar si valor numérico distinto de cero o si es texto no vacío significativo
                if (num is not None and abs(num) > 0) or (isinstance(val, str) and val.strip() not in ['', '-', '0']):
                    cc_count += 1
        else:
            vistos = set()
            for nombre, col in cc_indices_conocidos.items():
                if nombre == 'EB' and 'PS' in cc_indices_conocidos:
                    continue
                val = row[col] if col < len(row) else None
                num = _parse_numeric(val)
                if ((num is not None and abs(num) > 0) or (isinstance(val, str) and val and val.strip() not in ['', '-', '0'])) and nombre not in vistos:
                    cc_count += 1
                    vistos.add(nombre)

        clave = f"{tipo_doc_nombre} con {cc_count}CC"
        grupos[clave] = grupos.get(clave, 0) + 1
        grupos_filas.setdefault(clave, []).append((tipo_doc, tipo_doc_nombre, cc_count, row_idx))
        filas_originales[row_idx] = row
        debug_filas.append({
            'fila_excel': row_idx,
            'tipo_doc': tipo_doc,
            'tipo_doc_nombre': tipo_doc_nombre,
            'cc_count': cc_count,
            'clave': clave
        })

    wb_in.close()

    # Crear Excel de salida
    wb_out = Workbook()
    default_sheet = wb_out.active
    wb_out.remove(default_sheet)
    headers_salida = get_headers_salida_contabilidad()

    # Asegurar incorporación de la nueva columna requerida si aún no existe.
    # Se intenta colocar inmediatamente después de 'Monto 3 Detalle Libro' para mantener coherencia.
    if 'Monto Suma Detalle Libro' not in headers_salida:
        if 'Monto 3 Detalle Libro' in headers_salida:
            idx_m3 = headers_salida.index('Monto 3 Detalle Libro')
            headers_salida.insert(idx_m3 + 1, 'Monto Suma Detalle Libro')
        else:
            headers_salida.append('Monto Suma Detalle Libro')

    def sanitize(name: str) -> str:
        s = str(name).replace(':', '-').replace('/', '-').replace('\\', '-')
        return s[:31] if len(s) > 31 else s

    excel_json = {'sheets': []}

    for clave in sorted(grupos.keys()):
        sheet_name = sanitize(clave)
        ws = wb_out.create_sheet(title=sheet_name)
        for col_idx, h in enumerate(headers_salida, start=1):
            ws.cell(row=1, column=col_idx, value=h)
        row_cursor = 2
        row_counter = 1  # Contador secuencial para columna NUMERO

        filas_grupo = grupos_filas.get(clave, [])
        # Mapeo header salida -> índice para escribir
        header_to_col = {h: i + 1 for i, h in enumerate(headers_salida)}
        sheet_json = {
            'nombre': sheet_name,
            'headers': headers_salida,
            'rows': [],
        }
        excel_json['sheets'].append(sheet_json)

        # Detectar índices de columnas relevantes en entrada (heurística case-insensitive)
        headers_lower = [str(h).strip().lower() for h in headers]
        def find_idx(posibles):
            for nombre in headers_lower:
                if nombre in posibles:
                    return headers_lower.index(nombre)
            return None

        idx_monto_neto = find_idx({'monto neto', 'neto', 'monto_neto'})
        idx_monto_iva_rec = find_idx({'monto iva recuperable', 'iva recuperable', 'monto iva', 'iva'})
        idx_monto_total = find_idx({'monto total', 'total', 'monto_total'})
        # Funciones auxiliares
        def trunc(v):
            try:
                return int(float(v))
            except Exception:
                return 0

        for idx_fila, (tipo_doc_str, tipo_doc_nombre, cc_count, fila_original_idx) in enumerate(filas_grupo, start=1):
            row_in = filas_originales.get(fila_original_idx, [])
            monto_neto = _parse_numeric(row_in[idx_monto_neto]) if idx_monto_neto is not None and idx_monto_neto < len(row_in) else 0.0
            if monto_neto is None:
                monto_neto = 0.0

            tipo_doc_valor = tipo_doc_nombre or tipo_doc_str

            # Issue #174: Para tipo doc 33, sumar monto exento al monto neto
            monto_exento = 0.0
            if tipo_doc_str == '33' and idx_monto_exento is not None:
                monto_exento = _parse_numeric(row_in[idx_monto_exento]) if idx_monto_exento < len(row_in) else 0.0
                if monto_exento is None:
                    monto_exento = 0.0
                # No sumamos monto_exento a monto_neto - van separados

            # Issues #3 y #4: Extraer folio para mapeo a Nro. Docto. Conciliación
            folio = ""
            if idx_folio is not None and idx_folio < len(row_in):
                folio_val = row_in[idx_folio]
                folio = str(folio_val) if folio_val is not None else ""

            # Extraer RUT Proveedor para mapeo a "Codigo Auxiliar"
            rut_proveedor = ""
            if idx_rut_proveedor is not None and idx_rut_proveedor < len(row_in):
                rut_val = row_in[idx_rut_proveedor]
                if rut_val is not None:
                    rut_str = str(rut_val)
                    # Limpiar RUT: quitar guión y dígito verificador
                    if '-' in rut_str:
                        rut_proveedor = rut_str.split('-')[0]  # Solo la parte antes del guión
                    else:
                        rut_proveedor = rut_str
                else:
                    rut_proveedor = ""

            # Extraer Fecha Docto para mapeo a "Fecha Emisión" y "Fecha Vencimiento"
            fecha_docto = ""
            if idx_fecha_docto is not None and idx_fecha_docto < len(row_in):
                fecha_val = row_in[idx_fecha_docto]
                if fecha_val is not None:
                    # Si es una fecha de Excel (datetime), formatear como DD/MM/AAAA
                    try:
                        from datetime import datetime
                        if isinstance(fecha_val, datetime):
                            fecha_docto = fecha_val.strftime("%d/%m/%Y")
                        else:
                            # Si es string, mantenerlo como está
                            fecha_docto = str(fecha_val)
                    except:
                        fecha_docto = str(fecha_val) if fecha_val is not None else ""

            monto_total_input = _parse_numeric(row_in[idx_monto_total]) if idx_monto_total is not None and idx_monto_total < len(row_in) else None
            monto_iva_rec_input = _parse_numeric(row_in[idx_monto_iva_rec]) if idx_monto_iva_rec is not None and idx_monto_iva_rec < len(row_in) else None
            # IVA: si no existe columna o valor, calcular 0.19 * neto (truncado)
            iva_monto = monto_iva_rec_input if (monto_iva_rec_input is not None) else trunc(monto_neto * 0.19)
            # Total: si no existe usar neto + iva
            if monto_total_input is None:
                monto_total = (monto_neto + (iva_monto if monto_iva_rec_input is None else iva_monto))
            else:
                monto_total = monto_total_input

            # Recolectar montos por CC (cada fila de gasto corresponde a 1 CC)
            # Los gastos se calculan solo sobre el monto neto, monto exento va separado
            # PERO en "Monto al Debe Moneda Base" se suma neto + exento
            base_calculo_gastos = monto_neto
            base_debe_moneda_base = monto_neto + monto_exento  # Para columna "Monto al Debe Moneda Base"
            
            gastos_rows = []  # lista de (descripcion, debe_detalle, debe_moneda_base, codigo_cc)
            if cc_count > 0:
                if cc_start is not None and cc_end is not None:
                    for col in range(cc_start, cc_end):
                        if col >= len(row_in):
                            continue
                        val = row_in[col]
                        perc = _parse_numeric(val)
                        if perc is None:
                            continue
                        if abs(perc) > 0:
                            debe_detalle = (perc / 100.0) * base_calculo_gastos  # Para los Monto Detalle
                            debe_moneda_base = (perc / 100.0) * base_debe_moneda_base  # Para Monto al Debe Moneda Base
                            codigo_cc = headers[col] if col < len(headers) else f'CC{col}'
                            gastos_rows.append((f'Gasto {codigo_cc}', debe_detalle, debe_moneda_base, codigo_cc))
                else:
                    for nombre, col in cc_indices_conocidos.items():
                        if col >= len(row_in):
                            continue
                        val = row_in[col]
                        perc = _parse_numeric(val)
                        if perc is None:
                            continue
                        if abs(perc) > 0:
                            debe_detalle = (perc / 100.0) * base_calculo_gastos  # Para los Monto Detalle
                            debe_moneda_base = (perc / 100.0) * base_debe_moneda_base  # Para Monto al Debe Moneda Base
                            gastos_rows.append((f'Gasto {nombre}', debe_detalle, debe_moneda_base, nombre))

            suma_debe_gastos = sum(g[1] for g in gastos_rows)  # Suma de debe_detalle (sin exento)

            def _truncate_number(v):
                """Trunca (corta decimales) hacia cero cualquier número convertible a float."""
                if v is None:
                    return None
                try:
                    return int(float(v))
                except Exception:
                    return v

            def write_row(descripcion, debe=None, haber=None, extra=None):
                nonlocal row_cursor, row_counter
                row_values = [None] * len(headers_salida)
                
                # Escribir contador secuencial en columna NUMERO (primera columna)
                col_numero = header_to_col.get('Numero', 1)
                ws.cell(row=row_cursor, column=col_numero, value=row_counter)
                row_values[col_numero - 1] = row_counter
                row_counter += 1
                
                if debe is not None:
                    valor = _truncate_number(debe)
                    col = header_to_col.get('Monto al Debe Moneda Base', 3)
                    ws.cell(row=row_cursor, column=col, value=valor)
                    row_values[col - 1] = valor
                if haber is not None:
                    valor = _truncate_number(haber)
                    col = header_to_col.get('Monto al Haber Moneda Base', 4)
                    ws.cell(row=row_cursor, column=col, value=valor)
                    row_values[col - 1] = valor
                col_desc = header_to_col.get('Descripción Movimiento', 5)
                ws.cell(row=row_cursor, column=col_desc, value=descripcion)
                row_values[col_desc - 1] = descripcion
                if extra:
                    for hname, val in extra.items():
                        col_idx_h = header_to_col.get(hname)
                        if col_idx_h:
                            valor = _truncate_number(val)
                            ws.cell(row=row_cursor, column=col_idx_h, value=valor)
                            row_values[col_idx_h - 1] = valor
                sheet_json['rows'].append(row_values)
                row_cursor += 1

            # Tipos 33 / 64: IVA + Gastos + Proveedor (Proveedor al final)
            if tipo_doc_str in ['33', '64']:
                # Fila IVA
                iva_truncado = trunc(iva_monto)
                write_row(
                    descripcion=f'IVA Doc {fila_original_idx}',
                    debe=iva_truncado,
                    haber=None,
                    extra={
                        'Código Plan de Cuenta': cuentas_globales.get('iva'),
                        'Fecha Emisión Docto.(DD/MM/AAAA)': fecha_docto,
                        'Fecha Vencimiento Docto.(DD/MM/AAAA)': fecha_docto,
                        'Tipo Documento': tipo_doc_valor,
                        'Numero Doc': folio
                    }
                )
                
                # Filas de Gasto - Truncar cada una y ajustar la última para cuadratura
                gastos_truncados = []
                suma_gastos_truncados = 0
                
                for idx, (desc_gasto, debe_detalle, debe_moneda_base, codigo_cc) in enumerate(gastos_rows):
                    gasto_truncado = trunc(debe_moneda_base)
                    gastos_truncados.append((desc_gasto, debe_detalle, gasto_truncado, codigo_cc))
                    suma_gastos_truncados += gasto_truncado
                
                # Ajuste de cuadratura: modificar el último gasto
                # Debe total esperado = monto_total - iva
                # Haber = monto_total (en Proveedor)
                if gastos_truncados:
                    debe_esperado = trunc(monto_total) - iva_truncado
                    diferencia = debe_esperado - suma_gastos_truncados
                    
                    # Ajustar el último gasto
                    ultimo_gasto = list(gastos_truncados[-1])
                    ultimo_gasto[2] = ultimo_gasto[2] + diferencia  # Ajustar monto truncado
                    gastos_truncados[-1] = tuple(ultimo_gasto)
                
                # Escribir gastos ajustados
                for desc_gasto, debe_detalle, gasto_final, codigo_cc in gastos_truncados:
                    codigo_cc_final = mapeo_cc_param.get(codigo_cc, codigo_cc)
                    write_row(
                        descripcion=desc_gasto,
                        debe=gasto_final,
                        haber=None,
                        extra={
                            'Código Centro de Costo': codigo_cc_final,
                            'Código Plan de Cuenta': cuentas_globales.get('gasto_default'),
                            'Fecha Emisión Docto.(DD/MM/AAAA)': fecha_docto,
                            'Fecha Vencimiento Docto.(DD/MM/AAAA)': fecha_docto,
                            
                            'Tipo Documento': tipo_doc_valor,
                            'Numero Doc': folio
                        }
                    )
                
                # Fila Proveedores AL FINAL (usa suma ajustada)
                suma_debe_final = sum(g[2] for g in gastos_truncados) if gastos_truncados else 0
                monto1 = suma_debe_final
                monto2 = monto_exento if tipo_doc_str == '33' else 0.0
                monto3 = iva_truncado
                
                write_row(
                    descripcion=f'Proveedor Doc {fila_original_idx}',
                    debe=None,
                    haber=trunc(monto_total),
                    extra={
                        'Monto 1 Detalle Libro': monto1,
                        'Monto 2 Detalle Libro': trunc(monto2) if monto2 else None,
                        'Monto 3 Detalle Libro': monto3,
                        'Monto Suma Detalle Libro': monto1 + (trunc(monto2) if monto2 else 0) + monto3,
                        'Código Plan de Cuenta': cuentas_globales.get('proveedores'),
                        'Codigo Auxiliar': rut_proveedor,
                        'Fecha Emisión Docto.(DD/MM/AAAA)': fecha_docto,
                        'Fecha Vencimiento Docto.(DD/MM/AAAA)': fecha_docto,
                        
                        'Tipo Documento': tipo_doc_valor,
                        'Numero Doc': folio
                    }
                )
            elif tipo_doc_str == '34':
                # Tipo 34 (exento): Gastos + Proveedor (Proveedor al final)
                # Truncar gastos y ajustar último para cuadratura
                gastos_truncados = []
                suma_gastos_truncados = 0
                
                for idx, (desc_gasto, debe_detalle, debe_moneda_base, codigo_cc) in enumerate(gastos_rows):
                    gasto_truncado = trunc(debe_moneda_base)
                    gastos_truncados.append((desc_gasto, debe_detalle, gasto_truncado, codigo_cc))
                    suma_gastos_truncados += gasto_truncado
                
                # Ajuste de cuadratura
                if gastos_truncados:
                    debe_esperado = trunc(monto_total) if monto_total is not None else suma_gastos_truncados
                    diferencia = debe_esperado - suma_gastos_truncados
                    
                    # Ajustar el último gasto
                    ultimo_gasto = list(gastos_truncados[-1])
                    ultimo_gasto[2] = ultimo_gasto[2] + diferencia
                    gastos_truncados[-1] = tuple(ultimo_gasto)
                
                # Escribir gastos ajustados
                for desc_gasto, debe_detalle, gasto_final, codigo_cc in gastos_truncados:
                    codigo_cc_final = mapeo_cc_param.get(codigo_cc, codigo_cc)
                    write_row(
                        descripcion=desc_gasto,
                        debe=gasto_final,
                        haber=None,
                        extra={
                            'Código Centro de Costo': codigo_cc_final,
                            'Código Plan de Cuenta': cuentas_globales.get('gasto_default'),
                            'Fecha Emisión Docto.(DD/MM/AAAA)': fecha_docto,
                            'Fecha Vencimiento Docto.(DD/MM/AAAA)': fecha_docto,
                            
                            'Tipo Documento': tipo_doc_valor,
                            'Numero Doc': folio
                        }
                    )
                
                # Fila Proveedores AL FINAL
                suma_debe_final = sum(g[2] for g in gastos_truncados) if gastos_truncados else 0
                monto_total_truncado = trunc(monto_total) if monto_total is not None else suma_debe_final
                
                write_row(
                    descripcion=f'Proveedor Doc {fila_original_idx}',
                    debe=None,
                    haber=monto_total_truncado,
                    extra={
                        'Monto 2 Detalle Libro': suma_debe_final,
                        'Monto 3 Detalle Libro': None,
                        'Monto Suma Detalle Libro': suma_debe_final,
                        'Código Plan de Cuenta': cuentas_globales.get('proveedores'),
                        'Codigo Auxiliar': rut_proveedor,
                        'Fecha Emisión Docto.(DD/MM/AAAA)': fecha_docto,
                        'Fecha Vencimiento Docto.(DD/MM/AAAA)': fecha_docto,
                        
                        'Tipo Documento': tipo_doc_valor,
                        'Numero Doc': folio
                    }
                )
            elif tipo_doc_str == 'COMO':
                # Tipo COMO: Gastos + Proveedor (sin Monto Detalle)
                # Truncar gastos y ajustar último para cuadratura
                gastos_truncados = []
                suma_gastos_truncados = 0
                
                for idx, (desc_gasto, debe_detalle, debe_moneda_base, codigo_cc) in enumerate(gastos_rows):
                    gasto_truncado = trunc(debe_moneda_base)
                    gastos_truncados.append((desc_gasto, debe_detalle, gasto_truncado, codigo_cc))
                    suma_gastos_truncados += gasto_truncado
                
                # Ajuste de cuadratura
                if gastos_truncados:
                    debe_esperado = trunc(monto_total) if monto_total is not None else suma_gastos_truncados
                    diferencia = debe_esperado - suma_gastos_truncados
                    
                    # Ajustar el último gasto
                    ultimo_gasto = list(gastos_truncados[-1])
                    ultimo_gasto[2] = ultimo_gasto[2] + diferencia
                    gastos_truncados[-1] = tuple(ultimo_gasto)
                
                # Escribir gastos ajustados
                for desc_gasto, debe_detalle, gasto_final, codigo_cc in gastos_truncados:
                    codigo_cc_final = mapeo_cc_param.get(codigo_cc, codigo_cc)
                    write_row(
                        descripcion=desc_gasto,
                        debe=gasto_final,
                        haber=None,
                        extra={
                            'Código Centro de Costo': codigo_cc_final,
                            'Código Plan de Cuenta': cuentas_globales.get('gasto_default'),
                            'Fecha Emisión Docto.(DD/MM/AAAA)': fecha_docto,
                            'Fecha Vencimiento Docto.(DD/MM/AAAA)': fecha_docto,
                            
                            'Tipo Documento': tipo_doc_valor,
                            'Numero Doc': folio
                        }
                    )
                
                # Fila Proveedores AL FINAL (sin campos Monto Detalle)
                suma_debe_final = sum(g[2] for g in gastos_truncados) if gastos_truncados else 0
                monto_total_truncado = trunc(monto_total) if monto_total is not None else suma_debe_final
                
                write_row(
                    descripcion=f'Proveedor Doc {fila_original_idx}',
                    debe=None,
                    haber=monto_total_truncado,
                    extra={
                        'Código Plan de Cuenta': cuentas_globales.get('proveedores'),
                        'Codigo Auxiliar': rut_proveedor,
                        'Fecha Emisión Docto.(DD/MM/AAAA)': fecha_docto,
                        'Fecha Vencimiento Docto.(DD/MM/AAAA)': fecha_docto,
                        'Tipo Documento': tipo_doc_valor,
                        'Numero Doc': folio
                    }
                )
            elif tipo_doc_str == '61':
                # Tipo 61: espejo invertido de 33 (Nota de Crédito)
                # En 33: IVA (Debe), Gastos (Debe), Proveedor (Haber)
                # En 61: IVA (Haber), Gastos (Haber), Proveedor (Debe)
                
                # Fila IVA (invertido -> Haber)
                iva_truncado = trunc(iva_monto)
                write_row(
                    descripcion=f'IVA Doc {fila_original_idx}',
                    debe=None,
                    haber=iva_truncado,
                    extra={
                        'Código Plan de Cuenta': cuentas_globales.get('iva'),
                        'Fecha Emisión Docto.(DD/MM/AAAA)': fecha_docto,
                        'Fecha Vencimiento Docto.(DD/MM/AAAA)': fecha_docto,
                        
                        'Tipo Docto. Conciliación': tipo_doc_valor,
                        'Nro. Docto. Conciliación': folio
                    }
                )
                
                # Filas Gasto (invertidas -> Haber) - Truncar y ajustar última
                gastos_truncados = []
                suma_gastos_truncados = 0
                
                for idx, (desc_gasto, debe_detalle, debe_moneda_base, codigo_cc) in enumerate(gastos_rows):
                    gasto_truncado = trunc(debe_moneda_base)
                    gastos_truncados.append((desc_gasto, debe_detalle, gasto_truncado, codigo_cc))
                    suma_gastos_truncados += gasto_truncado
                
                # Ajuste de cuadratura: Haber = monto_total (IVA + Gastos)
                if gastos_truncados:
                    haber_esperado = trunc(monto_total) - iva_truncado
                    diferencia = haber_esperado - suma_gastos_truncados
                    
                    # Ajustar el último gasto
                    ultimo_gasto = list(gastos_truncados[-1])
                    ultimo_gasto[2] = ultimo_gasto[2] + diferencia
                    gastos_truncados[-1] = tuple(ultimo_gasto)
                
                # Escribir gastos ajustados (al Haber)
                for desc_gasto, debe_detalle, gasto_final, codigo_cc in gastos_truncados:
                    codigo_cc_final = mapeo_cc_param.get(codigo_cc, codigo_cc)
                    write_row(
                        descripcion=desc_gasto,
                        debe=None,
                        haber=gasto_final,
                        extra={
                            'Código Centro de Costo': codigo_cc_final,
                            'Código Plan de Cuenta': cuentas_globales.get('gasto_default'),
                            'Fecha Emisión Docto.(DD/MM/AAAA)': fecha_docto,
                            'Fecha Vencimiento Docto.(DD/MM/AAAA)': fecha_docto,
                            
                            'Tipo Documento': tipo_doc_valor,
                            'Numero Doc': folio
                        }
                    )
                
                # Fila Proveedores AL FINAL (invertido -> Debe)
                suma_haber_final = sum(g[2] for g in gastos_truncados) if gastos_truncados else 0
                monto1 = suma_haber_final
                monto2 = 0.0  # Notas de crédito no llevan monto exento
                monto3 = iva_truncado
                
                write_row(
                    descripcion=f'Proveedor Doc {fila_original_idx}',
                    debe=trunc(monto_total),
                    haber=None,
                    extra={
                        'Monto 1 Detalle Libro': monto1,
                        'Monto 2 Detalle Libro': None,
                        'Monto 3 Detalle Libro': monto3,
                        'Monto Suma Detalle Libro': monto1 + monto3,
                        'Código Plan de Cuenta': cuentas_globales.get('proveedores'),
                        'Codigo Auxiliar': rut_proveedor,
                        'Fecha Emisión Docto.(DD/MM/AAAA)': fecha_docto,
                        'Fecha Vencimiento Docto.(DD/MM/AAAA)': fecha_docto,
                        'Tipo Documento': tipo_doc_valor,
                        'Numero Doc': folio
                    }
                )
            else:
                # Tipos desconocidos: de momento no se generan movimientos (queda hoja vacía)
                pass

    buffer = BytesIO()
    wb_out.save(buffer)
    excel_content = buffer.getvalue()

    # Guardar Excel en Redis binario
    redis_client_bin = get_redis_client_db1_binary()
    redis_client_bin.setex(f"rg_step1_excel:{usuario_id}:{task_id}", 300, excel_content)

    # Actualizar meta
    metadata.update({
        'estado': 'completado',
        'fin': timezone.now().isoformat(),
        'total_filas': total_filas,
        'grupos': list(sorted(grupos.keys())),
        'archivo_excel_disponible': True,
        'debug_filas': debug_filas[:200]  # limitar tamaño
    })
    redis_client.setex(
        f"rg_step1_meta:{usuario_id}:{task_id}", 300, json.dumps(metadata, ensure_ascii=False)
    )

    if rendicion:
        try:
            rendicion.datos_archivo = {
                'archivo_nombre': archivo_nombre,
                'task_id': task_id,
                'parametros_contables': parametros_contables,
                'metadata': metadata,
                'excel_json': excel_json,
                'excel_tamano_bytes': len(excel_content),
                'excel_content_type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            }
            rendicion.save(update_fields=['datos_archivo', 'updated_at'])
        except Exception as e:
            print(f"Error actualizando datos_archivo en Rendicion {getattr(rendicion, 'id', 'desconocida')}: {str(e)}")

    return {
        'task_id': task_id,
        'estado': 'completado',
        'total_filas': total_filas,
        'total_grupos': len(grupos),
        'archivo_excel_disponible': True,
    }
